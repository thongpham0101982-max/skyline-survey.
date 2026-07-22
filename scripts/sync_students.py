import os
import re
import sys
import requests
from openpyxl import load_workbook
from colorama import init, Fore, Style

init(autoreset=True)

WEBSITE_API_URL = "http://localhost:3000/api/sync-excel-students"
API_SECRET_TOKEN = "skyline_sync_secret_token_2026"

def log_info(msg): print(f"{Fore.CYAN}[INFO]{Style.RESET_ALL} {msg}")
def log_success(msg): print(f"{Fore.GREEN}[SUCCESS]{Style.RESET_ALL} {msg}")
def log_warning(msg): print(f"{Fore.YELLOW}[WARNING]{Style.RESET_ALL} {msg}")
def log_error(msg): print(f"{Fore.RED}[ERROR]{Style.RESET_ALL} {msg}")

def parse_excel_and_sync(file_path):
    if not os.path.exists(file_path):
        log_error(f"Không tìm thấy file: {file_path}")
        return

    log_info(f"Đang phân tích file: {os.path.basename(file_path)}...")
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active

    academic_year = ""
    class_name = ""
    students_data = []

    for r in range(1, 9):
        for c in range(1, 10):
            cell_value = str(sheet.cell(row=r, column=c).value or "").strip()
            if "năm học" in cell_value.lower():
                match = re.search(r"năm học:\s*([\d\s\-–]+)", cell_value, re.IGNORECASE)
                if match:
                    academic_year = match.group(1).replace(" ", "").strip()
            if "lớp" in cell_value.lower():
                match = re.search(r"lớp:\s*([^\n\r]+)", cell_value, re.IGNORECASE)
                if match:
                    class_name = match.group(1).strip()

    if not academic_year or not class_name:
        log_error("Không thể tự động nhận diện \"Năm học\" hoặc \"Lớp\" trong tiêu đề file Excel.")
        return

    log_success(f"Nhận diện thành công -> Năm học: {academic_year} | Lớp: {class_name}")

    for r in range(9, sheet.max_row + 1):
        stt = sheet.cell(row=r, column=1).value
        student_code = sheet.cell(row=r, column=2).value
        student_name = sheet.cell(row=r, column=3).value

        if not stt or not student_code or not student_name:
            continue

        gender_val = sheet.cell(row=r, column=5).value
        is_female = str(gender_val).strip() in ["✓", "x", "1", "True"]

        dob_val = sheet.cell(row=r, column=4).value
        dob_str = ""
        if dob_val:
            if hasattr(dob_val, "strftime"):
                dob_str = dob_val.strftime("%Y-%m-%d")
            else:
                dob_str = str(dob_val).strip()

        student = {
            "studentCode": str(student_code).strip(),
            "studentName": str(student_name).strip(),
            "gender": "Nữ" if is_female else "Nam",
            "dateOfBirth": dob_str
        }
        students_data.append(student)

    if not students_data:
        log_warning("Không tìm thấy học sinh nào hợp lệ để đồng bộ.")
        return

    log_info(f"Đã phân tích xong {len(students_data)} học sinh. Đang tiến hành đồng bộ qua API...")

    payload = {
        "academicYearName": academic_year,
        "className": class_name,
        "students": students_data
    }
    
    headers = {
        "Authorization": f"Bearer {API_SECRET_TOKEN}",
        "Content-Type": "application/json"
    }

    try:
        response = requests.post(WEBSITE_API_URL, json=payload, headers=headers, timeout=30)
        if response.status_code == 200:
            result = response.json()
            log_success(f"ĐỒNG BỘ THÀNH CÔNG! Đã cập nhật/thêm mới: {result.get(\"count\", 0)} hồ sơ học sinh.")
        else:
            log_error(f"Đồng bộ thất bại. Mã lỗi Server: {response.status_code}")
            log_error(f"Chi tiết: {response.text}")
    except Exception as e:
        log_error(f"Lỗi kết nối mạng đến server: {str(e)}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Sử dụng: python sync_students.py <duong_dan_file_excel>")
        sys.exit(1)
        
    excel_file = sys.argv[1]
    parse_excel_and_sync(excel_file)
